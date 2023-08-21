import datetime
import os.path
import random
import time
from selenium.webdriver.support import expected_conditions as EC
import pyperclip
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tqdm import tqdm
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from bs4 import BeautifulSoup
import os
import requests
from web3 import Web3
import xlsxwriter
from openpyxl import load_workbook


class OpenSeaParser:
    def __init__(self):
        self.url = input('Введите ссылку на коллекцию: ')
        self.name_folder = input('Введите название папки: ')
        self.with_twitter = input('Введите название таблицы с twitter: ')
        self.no_twitter = input('Введите название таблицы без twitter: ')

        self.create_table()

        self.firefox_options = FirefoxOptions()
        self.firefox_options.add_argument("--disable-blink-features=AutomationControlled")
        self.firefox_options.add_argument("--disable-extensions")
        self.firefox_options.add_argument("--disable-gpu")
        self.firefox_options.add_argument("--disable-dev-shm-usage")
        self.firefox_options.add_argument("--no-sandbox")
        self.firefox_options.add_argument("--disable-features=VizDisplayCompositor")
        self.firefox_options.set_preference("excludeSwitches", "enable-automation")
        self.firefox_options.set_preference('useAutomationExtension', False)
        self.firefox_options.set_preference('permissions.default.image', 2)
        self.firefox_options.set_preference('permissions.default.video', 2)
        self.firefox_options.set_preference('dom.ipc.plugins.enabled.libflashplayer.so', 'false')
        self.firefox_options.set_preference('network.proxy.type', 0)

        self.user_agent = [
            'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.42',
            'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.42',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.42',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.42',
            'Mozilla/5.0 (Windows NT 6.1; rv:109.0) Gecko/20100101 Firefox/114.0',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:109.0) Gecko/20100101 Firefox/114.0',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:109.0) Gecko/20100101 Firefox/114.0',
            'Mozilla/5.0 (Windows NT 6.3; rv:109.0) Gecko/20100101 Firefox/114.0',
            'Mozilla/5.0 (Windows NT 6.3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 OPR/98.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 OPR/98.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 OPR/97.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 OPR/97.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 OPR/97.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 OPR/96.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 OPR/96.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 OPR/96.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 OPR/95.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 OPR/95.0.0.0',
            'Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0',
            'Mozilla/5.0 (Windows NT 10.0; rv:109.0) Gecko/20100101 Firefox/112.0',
            'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:109.0) Gecko/20100101 Firefox/112.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0',
            'Mozilla/5.0 (Windows NT 6.1; rv:109.0) Gecko/20100101 Firefox/111.0',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:109.0) Gecko/20100101 Firefox/111.0'
        ]

        self.driver = webdriver.Firefox(options=self.firefox_options)

        self.url_to_user = os.path.join(os.getcwd(), 'temporarily', 'url_to_user.txt')
        self.url_to_profile = os.path.join(os.getcwd(), 'temporarily', 'url_to_profile.txt')
        self.miss_profile = os.path.join(os.getcwd(), 'temporarily', 'miss_profile.txt')

    def create_table(self):
        self.create_table_(
            name_folder=self.name_folder,
            with_twitter=self.with_twitter,
            no_twitter=self.no_twitter
        )

    def create_table_(self, name_folder: str, with_twitter: str, no_twitter: str):
        os.makedirs(f'./{name_folder}')
        workbook = xlsxwriter.Workbook(f'./{name_folder}/{with_twitter}.xlsx')
        worksheet = workbook.add_worksheet()

        column_sizes = {
            'Name': 165,
            'Twitter': 320,
            'Wallet': 320,
            'Debank': 65,
            'Offers': 65,
            'Description': 770
        }

        for col_num, (col_name, width_pixels) in enumerate(column_sizes.items(), start=1):
            worksheet.set_column(col_num, col_num, width_pixels / 7)

        bold_format = workbook.add_format({'bold': True})

        headers = ['Name', 'Twitter', 'Wallet', 'Debank', 'Offers', 'Description']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, bold_format)  # Используем жирный формат

        workbook.close()

        workbook = xlsxwriter.Workbook(f'./{name_folder}/{no_twitter}.xlsx')
        worksheet = workbook.add_worksheet()

        column_sizes = {
            'Wallet': 320
        }

        for col_num, (col_name, width_pixels) in enumerate(column_sizes.items(), start=1):
            worksheet.set_column(col_num, col_num, width_pixels / 7)

        bold_format = workbook.add_format({'bold': True})

        headers = ['Wallet']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, bold_format)  # Используем жирный формат

        workbook.close()

    def save_with(self, name_fold: str, with_twitter: str,
                  name: str, desc: str,
                  wallet: str, debank: str,
                  offers: str, twitter: str = None
                  ):
        workbook = load_workbook(f'./{name_fold}/{with_twitter}.xlsx')
        worksheet = workbook.active
        new_data = [
            [name, twitter, wallet, debank, offers, desc],
        ]
        next_row = worksheet.max_row + 1
        for row_data in new_data:
            for col_num, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=next_row, column=col_num, value=cell_value)
        workbook.save(f'./{name_fold}/{with_twitter}.xlsx')

    def save_no_with(self, name_fold: str, no_twitter: str,
                     wallet: str
                     ):
        workbook = load_workbook(f'./{name_fold}/{no_twitter}.xlsx')
        worksheet = workbook.active
        new_data = [
            [wallet],
        ]
        next_row = worksheet.max_row + 1
        for row_data in new_data:
            for col_num, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=next_row, column=col_num, value=cell_value)
        workbook.save(f'./{name_fold}/{no_twitter}.xlsx')

    def check_path(self):
        if not os.path.exists(self.url_to_user):
            with open(self.url_to_user, 'w') as file:
                pass
        if not os.path.exists(self.url_to_profile):
            with open(self.url_to_profile, 'w') as file:
                pass
        if not os.path.exists(self.miss_profile):
            with open(self.miss_profile, 'w') as file:
                pass

    def check_balance(self, wallet_address: str) -> str | int:
        web3 = Web3(Web3.HTTPProvider('https://eth-rpc.gateway.pokt.network'))

        checksum_address = Web3.to_checksum_address(wallet_address)
        balance_wei = web3.eth.get_balance(checksum_address)

        def convert_to_usd(wei_balance):
            try:
                response = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=ethereum&vs_currencies=usd")
                eth_to_usd_rate = response.json().get("ethereum", {}).get("usd", 0)
                usd_balance = wei_balance * eth_to_usd_rate * 10 ** -18
                return usd_balance
            except Exception as e:
                return None

        usd_balance = convert_to_usd(balance_wei)

        if usd_balance is not None:
            return f"{usd_balance:.2f}"
        else:
            return 0

    def parsing_nft_to_user(self):
        self.check_path()
        try:
            self.driver.get("https://nftinit.com/assets/ink-cell")
            time.sleep(10)
            wait = WebDriverWait(self.driver, 30)

            content = self.driver.page_source
            soup = BeautifulSoup(content, 'lxml')
            all_html = soup.find_all('tr', class_='bg-d-dark-70')
            page = soup.find('span', class_='h-6 relative z-0 inline-flex shadow-sm rounded-md').find('button',
                                                                                                      class_='-ml-px relative inline-flex items-center px-2 border border-gray-300 bg-white text-sm font-medium text-gray-500 hover:bg-gray-50 focus:z-10 focus:outline-none focus:ring-1 focus:ring-indigo-500 focus:border-indigo-500').get_text()

            with open(self.url_to_user, 'a') as file:
                for link_element in all_html:
                    link_href = link_element.find('a', target='_blank').get('href')
                    if link_href.startswith('https://opensea.io/assets/'):
                        file.write(link_href + '\n')

            count_ = 2
            page = page.split('/')[1]

            for count in tqdm(range(int(page)), desc="Parsing Pages", unit="page"):
                try:
                    self.firefox_options.set_preference("general.useragent.override", random.choice(self.user_agent))
                    url = f'https://nftinit.com/assets/ink-cell?page={count_}&'
                    count_ += 1

                    self.driver.get(url)
                    time.sleep(10)
                    wait = WebDriverWait(self.driver, 30)
                    tbody = wait.until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "tbody.bg-d-dark-90.divide-y.divide-gray-800")))

                    content = self.driver.page_source
                    soup = BeautifulSoup(content, 'lxml')
                    all_html = soup.find_all('tr', class_='bg-d-dark-70')

                    if all_html == []:
                        all_html = soup.find('tbody', class_='bg-d-dark-90 divide-y divide-gray-800').find_all('tr')

                    with open(self.url_to_user, 'a') as file:
                        for link_element in all_html:
                            link_href = link_element.find('a', target='_blank').get('href')
                            if link_href.startswith('https://opensea.io/assets/'):
                                file.write(link_href + '\n')

                except Exception as e:
                    continue
        except Exception as e:
            print("Произошла ошибка:", e)

        finally:
            now = datetime.datetime.now()
            print(f'{now.strftime("%Y-%m-%d %H:%M")} | Парсинг NFT страниц закончен начинаю парсить URL профилей')
            self.parsing_url_profile()

    def parsing_url_profile(self):
        try:
            unique_urls = set()
            with open(self.url_to_user, 'r') as file:
                lines = file.readlines()
                for urls in lines:
                    unique_urls.add(urls.strip())

            with open(self.url_to_user, 'w') as file:
                for unique_url in unique_urls:
                    file.write(unique_url + '\n')

            with tqdm(total=len(unique_urls), desc='Check URL profile', unit='count') as pbar:
                for href in unique_urls:
                    self.driver.get(href)
                    time.sleep(5)
                    content = self.driver.page_source
                    soup = BeautifulSoup(content, 'lxml')

                    account_link = soup.find('a',
                                             class_='sc-1f719d57-0 jDhJBQ sc-57ceaf99-0 irUTbw AccountLink--ellipsis-overflow')
                    if account_link:
                        profile = 'https://opensea.io' + account_link.get('href') + '/bids'
                        with open(self.url_to_profile, 'a') as file:
                            file.write(profile + '\n')
                    else:
                        with open(self.miss_profile, 'a') as file:
                            file.write(href + '\n')

                    self.driver.delete_all_cookies()
                    pbar.update(1)
            profile_set_miss = set()

            with open(self.miss_profile, 'r') as file:
                lines = file.readlines()
                for urlf in lines:
                    profile_set_miss.add(urlf.strip())

            if profile_set_miss is not None:
                now = datetime.datetime.now()
                print(f'{now.strftime("%Y-%m-%d %H:%M")} | Проверка пропущенных URL профилей')
                with tqdm(total=len(profile_set_miss)) as pbar:
                    for href in profile_set_miss:
                        self.driver.get(href)
                        time.sleep(5)

                        content = self.driver.page_source
                        soup = BeautifulSoup(content, 'lxml')

                        account_link = soup.find('a',
                                                 class_='sc-1f719d57-0 jDhJBQ sc-57ceaf99-0 irUTbw AccountLink--ellipsis-overflow')
                        if account_link:
                            profile = 'https://opensea.io' + account_link.get('href')
                            with open(self.url_to_profile, 'a') as file:
                                file.write(profile + '\n')

                            self.driver.delete_all_cookies()
                            pbar.update(1)

                        self.driver.delete_all_cookies()



        except Exception as e:
            print("Произошла ошибка:", e)

        finally:
            now = datetime.datetime.now()
            print(f'{now.strftime("%Y-%m-%d %H:%M")} | Парсинг URL профилей окончен начинаю проверку профилей')
            self.parsing_profile_finnaly()

    def parsing_profile_finnaly(self):
        try:

            unique_urls = set()
            with open(self.url_to_profile, 'r') as file:
                lines = file.readlines()
                for urlv in lines:
                    unique_urls.add(urlv.strip())

            with open(self.url_to_profile, 'w') as file:
                for unique_url in unique_urls:
                    file.write(unique_url + '\n')

            with tqdm(total=len(unique_urls), desc='Check profile data', unit='count') as pbar:
                for href in unique_urls:
                    try:
                        self.firefox_options.set_preference("general.useragent.override",
                                                            random.choice(self.user_agent))
                        self.driver.delete_all_cookies()
                        self.driver.get(href)
                        time.sleep(10)
                        response = self.driver.page_source
                        soup = BeautifulSoup(response, 'lxml')
                        if soup.find('div', class_='heading-icon warning-icon'):
                            time.sleep(40)
                        twitter = soup.find('div',
                                            class_='sc-57ceaf99-0 sc-630fc9ab-0 sc-62cc93cd-0 kHrhvD bNkKFC lcXppp').find_all(
                            'a', class_='sc-1f719d57-0 eiItIQ')
                        if twitter != []:
                            for url_twitter in twitter:
                                check_twitter = url_twitter.get('href')
                                if check_twitter.startswith('https://twitter.com'):
                                    twitter = check_twitter
                        else:
                            twitter = None
                        wallet = soup.find('button', class_='sc-b267fe84-0 cEtajt')
                        if wallet:
                            click = self.driver.find_element(by=By.XPATH,
                                                             value='//*[@id="main"]/div/div/div/div[4]/div/div/div/div/div/div[1]/div[1]/div/button[2]')
                            click.click()
                            time.sleep(1)
                            try:
                                try:
                                    try:
                                        button = self.driver.find_element(by=By.XPATH,
                                                                          value='/html/body/div[1]/div/main/div/div/div/div[4]/div/div/div/div/div/div[1]/div[2]/div/div/div/ul/li[2]/button/div[2]')
                                    except:
                                        button = self.driver.find_element(by=By.XPATH,
                                                                          value='/html/body/div[1]/div/main/div/div/div/div[4]/div/div/div/div/div/div[1]/div[2]/div/div/div/ul/li[2]/button/div[1]/div/div[2]')
                                except:
                                    button = self.driver.find_element(by=By.XPATH,
                                                                      value='/html/body/div[1]/div/main/div/div/div/div[4]/div/div/div/div/div/div[1]/div[2]/div/div/div/ul/li[2]/button')
                            except:
                                button = self.driver.find_element(by=By.XPATH,
                                                                  value='//*[@id="main"]/div/div/div/div[4]/div/div/div/div/div/div[1]/div/div/button[2]')

                            button.click()
                            wallet = pyperclip.paste()
                            if wallet.startswith('0x') is False:
                                print('asdf')
                                continue
                        else:
                            wallet = self.driver.find_element(by=By.XPATH,
                                                              value='//*[@id="main"]/div/div/div/div[4]/div/div/div/div/div/div[1]/div/div/button[2]')
                            wallet.click()
                            wallet = pyperclip.paste()

                        name = soup.find('h1', class_='sc-57ceaf99-0 sc-bgqQcB bwyaYZ dTwVlG').get_text()
                        description = soup.find('div', class_='sc-3e08d4e-8 hhfftw')
                        if description:
                            description = description.get_text()

                        amount_offers = 0

                        offer = soup.find('ul', class_='sc-57ceaf99-0 sc-4bba1e52-0 kHrhvD KiXEg')
                        try:
                            if offer:
                                data = []
                                links = soup.find_all('a', class_='sc-1f719d57-0 eiItIQ sc-57ceaf99-0 eLkFjw')
                                for link in links:
                                    href = link['href']
                                    price = link.find_next('span', class_='sc-57ceaf99-0 sc-bgqQcB fXQuSw GczNV').text
                                    data.append((href, price.replace('$', '').replace(',', '.')))

                                sorted_data = sorted(data, key=lambda x: float(x[1][1:]), reverse=True)

                                unique_links = {}
                                for href, price in sorted_data:
                                    if href not in unique_links:
                                        unique_links[href] = price

                                for href, price in unique_links.items():
                                    amount_offers = amount_offers + float(price)
                        except:
                            pass

                        debank_balance = self.check_balance(wallet_address=wallet)

                        if twitter is None:
                            self.save_no_with(name_fold=self.name_folder,
                                         no_twitter=self.no_twitter,
                                         wallet=wallet)
                        else:
                            self.save_with(name_fold=self.name_folder,
                                      with_twitter=self.with_twitter,
                                      twitter=twitter or 'No',
                                      wallet=wallet or 'No',
                                      debank=debank_balance or 'No',
                                      name=name or 'No',
                                      desc=description or 'No',
                                      offers=offer or 'No')

                        self.driver.delete_all_cookies()
                        pbar.update(1)
                    except Exception as e:
                        self.driver.delete_all_cookies()
                        continue

        except Exception as e:
            print("Произошла ошибка:", e)
        finally:
            for temp in [self.miss_profile, self.url_to_profile, self.url_to_user]:
                with open(temp, "w") as file:
                    file.truncate(0)

            self.driver.quit()


if __name__ == "__main__":
    parser = OpenSeaParser()
    parser.parsing_nft_to_user()
